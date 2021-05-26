import time

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains
import openpyxl

path = "/home/ravali/Downloads/badge_44_accounts.csv"#mostly it supports only xls,xlsx files.

#Read data from file
workbook = openpyxl.load_workbook(path)
sheet = workbook.active
#workbook.get_sheet_by_name("sheet1") - to get it with sheetname.
rows = sheet.max_row
cols = sheet.max_column

print(rows)
print(cols)

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value,end="      ")

    print()

#write data into file
path1 = "/home/ravali/Downloads/test_accounts.csv"#mostly it supports only xls,xlsx files.
workbook = openpyxl.load_workbook(path1)
sheet = workbook.active

for r in range(1,6):
    for c in range(1,3):
        sheet.cell(row=r,column=c).value == "Welcome"

workbook.save()
#---------------------------